import pandas as pd
import numpy as np # numpy を使用して数値計算や条件分岐を効率的に行います
import io
import os # ファイルパスの存在確認のために os モジュールをインポート
import datetime # 日付取得のために追加

# テスト用

# openpyxl.utils.get_column_letter を使用するためにインポート (pandasが内部で使用するopenpyxlに依存)
# 通常、pandasと共にインストールされていれば利用可能
try:
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Border, Side, Font, Color, Alignment
except ImportError:
    # openpyxl がないか、utils が見つからない場合のフォールバックやエラー処理
    def get_column_letter(idx): # 簡単なフォールバック (限定的)
        letter = ''
        while idx > 0:
            idx, remainder = divmod(idx - 1, 26)
            letter = chr(65 + remainder) + letter
        return letter
    # スタイル関連のクラスもフォールバックが必要になるが、
    # このスクリプトの主要機能には openpyxl が必須のため、
    # ここでのフォールバックは限定的とし、通常は ImportError が発生しない前提とする
    class Border: pass
    class Side: pass
    class Font: pass
    class Color: pass
    class Alignment: pass


def create_repacking_priority_list_from_excel(file_path_or_obj, sheet_name=0):
    """
    Excelファイル (.xlsx) を処理し、小分け生産チーム向けの作業優先度リストをExcelファイルで生成します。
    1行目にタイトル「日付(mm月dd日) 小分け作成メモ」(フォントサイズ14pt)が記述されます。
    商品は「充足率」昇順を第一優先としてソートされます。
    商品名に「◇」を含み、かつ「今日入荷（作成）」が0でなく、かつ商品名末尾が「東一」でない商品が対象です。
    「箱/こもの」列は出力されません。充足率は小数点以下1桁のパーセント表示になり、列幅が固定されます。
    ヘッダー行とデータセルには罫線が引かれ、D列「本日作成」は太字、F列「不足数」は太字・赤字になります。
    全ての行高は18.0、ヘッダー行(2行目)は中央揃え（横・縦）、データ行は中央揃え（縦）になります。
    フッターとして各種注釈が最終データ行の1行下からA列に順に記述されます。
    出力ファイル名は「日付(mmdd)_小分け作業の判断指標.xlsx」となります。

    Args:
        file_path_or_obj (str or UploadedFile): 入力ExcelファイルのパスまたはStreamlitのUploadedFileオブジェクト。
        sheet_name (str or int, optional): 読み込むシートの名前またはインデックス。
                                           デフォルトは0 (最初のシート)。

    Returns:
        tuple: (成功フラグ, メッセージ, Excelファイル名, Excelデータ)
               成功フラグ (bool): 処理が成功したかどうか。
               メッセージ (str): 処理結果のメッセージ。
               Excelファイル名 (str): 生成されたExcelファイルの名前 (成功時のみ)。
               Excelデータ (bytes): 生成されたExcelファイルのバイナリデータ (成功時のみ)。
    """
    available_cols = [] # エラーメッセージ用に利用可能な列名を保持
    try:
        # pandas.read_excel はファイルパスまたはファイルライクオブジェクトを受け取れる
        df = pd.read_excel(file_path_or_obj, header=1, sheet_name=sheet_name)
        
        df.columns = df.columns.str.strip()
        available_cols = df.columns.tolist()

        # --- Excelファイル内の列名定義 (これらの名前がファイルの2行目ヘッダーに存在することを確認) ---
        product_code_col_excel = "商品コード"
        product_name_col_excel = "商品名"
        prev_day_stock_col_excel = "昨日残"
        today_received_col_excel = "今日入荷（作成）" 
        delivery_qty_col_excel = "納品数"
        action_needed_col_excel = "集荷便から降ろす数/小分けしないと足りない数" 

        # --- 必要な入力列がDataFrameに存在するか確認 ---
        required_input_cols_map = {
            "商品コード(A列)": product_code_col_excel,
            "商品名(B列)": product_name_col_excel,
            "昨日残(C列)": prev_day_stock_col_excel,
            "今日入荷作成(D列)": today_received_col_excel,
            "納品数(E列)": delivery_qty_col_excel,
            "小分け不足数(K列)": action_needed_col_excel
        }
        
        missing_excel_cols = []
        for display_name, actual_col_name in required_input_cols_map.items():
            if actual_col_name not in df.columns:
                missing_excel_cols.append(f"{display_name} (想定ヘッダー名: '{actual_col_name}')")

        if missing_excel_cols:
            return False, (f"Excelファイルに必要な列が見つかりません: {', '.join(missing_excel_cols)}。\n"
                          f"Excelファイルの最初のシートの2行目のヘッダー名を確認してください。\n"
                          f"読み込まれたExcelヘッダー: {available_cols}"), None, None

        # --- データフィルタリング ---
        df_temp_filtered = df[df[product_name_col_excel].astype(str).str.contains('◇', na=False)].copy()
        if df_temp_filtered.empty:
            return True, "対象商品（商品名に「◇」を含む）が見つかりませんでした。", None, None
        
        df_temp_filtered = df_temp_filtered[~df_temp_filtered[product_name_col_excel].astype(str).str.endswith('東一', na=False)].copy()
        if df_temp_filtered.empty:
            return True, "対象商品（商品名に「◇」を含み、かつ末尾が「東一」でない）が見つかりませんでした。", None, None

        # --- 数値列を数値型に変換 ---
        numeric_cols_to_convert = [
            prev_day_stock_col_excel,
            today_received_col_excel, 
            delivery_qty_col_excel,
            action_needed_col_excel
        ]
        for col in numeric_cols_to_convert:
            if col in df_temp_filtered.columns: 
                 df_temp_filtered.loc[:, col] = pd.to_numeric(df_temp_filtered[col], errors='coerce').fillna(0)

        df_filtered = df_temp_filtered[df_temp_filtered[today_received_col_excel] != 0].copy()
        if df_filtered.empty:
            return True, "対象商品（商品名に「◇」を含み、末尾が「東一」でなく、かつ「今日入荷（作成）」が0でない）は見つかりませんでした。", None, None
            
        # --- 新規列の計算 (df_filtered に対して行う) ---
        df_filtered.loc[:, 'calculated_充足率'] = np.where(
            df_filtered[delivery_qty_col_excel] != 0,
            df_filtered[prev_day_stock_col_excel] / df_filtered[delivery_qty_col_excel],
            0
        )
        
        df_filtered.loc[:, 'calculated_E_K_ratio'] = np.where(
            df_filtered[action_needed_col_excel] > 0, 
            df_filtered[delivery_qty_col_excel] / df_filtered[action_needed_col_excel],
            -1 
        )

        # --- 並び替え (df_filtered に対して行う) ---
        df_sorted = df_filtered.sort_values(
            by=['calculated_充足率', action_needed_col_excel, prev_day_stock_col_excel, 'calculated_E_K_ratio'],
            ascending=[True, True, True, False] 
        )

        # --- 出力用DataFrame作成 ---
        output_df = pd.DataFrame()
        output_df['商品コード'] = df_sorted[product_code_col_excel]
        output_df['商品名'] = df_sorted[product_name_col_excel]
        output_df['昨日残'] = df_sorted[prev_day_stock_col_excel]
        output_df['本日作成'] = df_sorted[today_received_col_excel] 
        output_df['納品数'] = df_sorted[delivery_qty_col_excel]
        output_df['不足数'] = df_sorted[action_needed_col_excel] 
        output_df['充足率'] = df_sorted['calculated_充足率'] 
        
        # --- 出力ファイル名の生成 ---
        # 日付(mmdd)を取得
        current_date_mmdd_filename = datetime.datetime.now().strftime("%m%d")
        output_filename = f"{current_date_mmdd_filename}_小分け作業の判断指標.xlsx"
        
        excel_buffer = io.BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            output_df.to_excel(writer, index=False, sheet_name='作業優先リスト', startrow=1) 
            
            # --- Excelの書式設定 ---
            workbook  = writer.book
            worksheet = writer.sheets['作業優先リスト']

            # --- タイトル行の追加と設定 ---
            current_date_title = datetime.datetime.now().strftime("%m月%d日") 
            title_text = f"{current_date_title} 小分け作成メモ"
            worksheet['B1'] = title_text
            title_font = Font(bold=True, size=14) 
            worksheet['B1'].font = title_font
            worksheet['B1'].alignment = Alignment(vertical='center') 

            # スタイル定義
            thin_border_side = Side(border_style="thin", color="000000")
            thin_border = Border(left=thin_border_side, 
                                 right=thin_border_side, 
                                 top=thin_border_side, 
                                 bottom=thin_border_side)
            bold_font = Font(bold=True) 
            red_font_for_shortage = Font(bold=True, color="FF0000")
            data_row_vertical_alignment = Alignment(vertical='center') 
            header_row_center_alignment = Alignment(horizontal='center', vertical='center') 

            # 列幅の設定
            column_widths = [9.0, 37.0, 7.5, 9.0, 7.5, 7.5, 7.5] 
            for i, width in enumerate(column_widths):
                column_letter = get_column_letter(i + 1) # 1-indexed
                worksheet.column_dimensions[column_letter].width = width

            # 「本日作成」列と「不足数」列の列文字を取得
            try:
                honjitsu_sakusei_col_index = output_df.columns.get_loc('本日作成') + 1
                honjitsu_sakusei_col_letter = get_column_letter(honjitsu_sakusei_col_index)
            except KeyError:
                honjitsu_sakusei_col_letter = None
                print("警告: '本日作成' 列が見つからず、書式設定をスキップします。")
            
            try:
                shortage_col_index = output_df.columns.get_loc('不足数') + 1
                shortage_col_letter = get_column_letter(shortage_col_index)
            except KeyError:
                shortage_col_letter = None 
                print("警告: '不足数' 列が見つからず、書式設定をスキップします。")

            # 全データセルに書式を適用 (タイトル行、ヘッダー行、データ行)
            # 行の高さもここで設定
            for row_idx in range(1, worksheet.max_row + 1): # 1から始まる行インデックス
                worksheet.row_dimensions[row_idx].height = 18.0
                for cell in worksheet[row_idx]: # 特定の行のセルをイテレート
                    current_cell_col_letter = get_column_letter(cell.column)
                    
                    if cell.row == 1: # タイトル行 (1行目)
                        if cell.column == 2: # B列のみ (タイトルセル)
                             cell.alignment = Alignment(vertical='center') 
                        # タイトル行の他のセルには罫線を適用しない
                    elif cell.row == 2: # ヘッダー行 (2行目)
                        cell.border = thin_border # ヘッダー行に罫線を適用
                        cell.alignment = header_row_center_alignment
                        if honjitsu_sakusei_col_letter and current_cell_col_letter == honjitsu_sakusei_col_letter:
                            cell.font = bold_font 
                    else: # データ行 (3行目以降)
                        cell.border = thin_border # データ行に罫線を適用
                        cell.alignment = data_row_vertical_alignment
                        if honjitsu_sakusei_col_letter and current_cell_col_letter == honjitsu_sakusei_col_letter:
                            cell.font = bold_font 
                        elif shortage_col_letter and current_cell_col_letter == shortage_col_letter:
                            cell.font = red_font_for_shortage 
            
            # 「充足率」列 (G列) にパーセント表示形式を適用 (データは3行目から)
            try:
                column_index_percent = output_df.columns.get_loc('充足率') + 1 
                column_letter_percent = get_column_letter(column_index_percent)
                
                for row_num in range(3, worksheet.max_row + 1): # データ行は3行目から
                    cell = worksheet[f'{column_letter_percent}{row_num}']
                    cell.number_format = '0.0%' 
            except KeyError:
                print("警告: '充足率' 列が見つからず、パーセント書式を適用できませんでした。")
            except Exception as e:
                print(f"警告: パーセント書式の適用中にエラーが発生しました: {e}")

            # --- フッターの追加 ---
            if worksheet.max_row >= 2 : # ヘッダー行(2行目)が存在すれば、その下にデータがあるか、少なくともヘッダーはある
                # 1行目のフッター (充足率の説明)
                footer1_row_num = worksheet.max_row + 2 
                worksheet.row_dimensions[footer1_row_num].height = 18.0
                footer1_text = "※充足率＝「納品数」に対する「昨日残数」の割合（昨日残数÷納品数）"
                footer1_cell = worksheet[f'A{footer1_row_num}'] 
                footer1_cell.value = footer1_text
                footer1_cell.alignment = data_row_vertical_alignment

                # 2行目のフッター (東一商品の注釈)
                footer2_row_num = footer1_row_num + 1 
                worksheet.row_dimensions[footer2_row_num].height = 18.0
                footer2_text = "※「東一」用の商品名の記載はありませんが、該当商品の不足数には反映されています。" 
                footer2_cell = worksheet[f'A{footer2_row_num}'] 
                footer2_cell.value = footer2_text
                footer2_cell.alignment = data_row_vertical_alignment
            
        excel_data = excel_buffer.getvalue()

        return True, f"処理が完了しました。「{output_filename}」を確認してください。", output_filename, excel_data

    except FileNotFoundError: 
        return False, f"指定されたファイルが見つかりません: {file_path_or_obj}", None, None
    except pd.errors.EmptyDataError: 
        return False, "Excelファイルが空か、データが読み取れませんでした。", None, None
    except ValueError as e: 
        if "Worksheet" in str(e) and "not found" in str(e):
             return False, f"Excelファイルの最初のシートの読み込み中にエラーが発生しました。ファイルが破損しているか、形式が正しくない可能性があります。エラー: {e}", None, None
        return False, f"Excelファイルの読み込み中にエラーが発生しました (ValueError): {e}", None, None
    except KeyError as e:
        return False, (f"Excelファイルに必要な列が見つかりません: {e}。\n"
                      f"スクリプト内の列名定義と、Excelファイルの最初のシートの2行目のヘッダー名が一致しているか確認してください。\n"
                      f"読み込まれたExcelヘッダー: {available_cols}"), None, None
    except Exception as e:
        return False, f"処理中に予期せぬエラーが発生しました: {e}", None, None
